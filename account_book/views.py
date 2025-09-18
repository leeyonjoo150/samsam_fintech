from django.shortcuts import render, redirect
from dashboard.models import Transaction
from manage_account.models import Account, AccountBookCategory, TransactionAccount, TransactionCash
from django.contrib.auth import get_user_model
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.db.models import Sum, OuterRef, Subquery, Q
from django.http import JsonResponse
import json
from django.contrib.auth.decorators import login_required
from datetime import datetime, timedelta, date # Added date import explicitly
from django.db import transaction
import calendar # Added calendar import
from django.utils.timezone import make_aware
from django.utils.timezone import localtime
import openpyxl
from django.http import HttpResponse
from manage_account.models import TransactionAccount, TransactionCash
from django.http import JsonResponse
from .models import Category

User = get_user_model()

def get_categories_json(request):
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Authentication required'}, status=401)

    category_type = request.GET.get('type')
    categories = []
    if category_type in ['income', 'expense']:
        filtered_categories = AccountBookCategory.objects.filter(cat_kind=category_type).order_by('cat_type')
        categories = [{'id': category.id, 'name': category.cat_type} for category in filtered_categories]
    return JsonResponse({'categories': categories})


def home(request):
    categories = AccountBookCategory.objects.all()
    
    # Get year and month from GET parameters, default to current month
    year = request.GET.get('year')
    month = request.GET.get('month')

    now = timezone.now()
    current_year = int(year) if year else now.year
    current_month = int(month) if month else now.month

    # Filter TransactionCash for the current month
    cash_transactions = TransactionCash.objects.filter(
        use_date__year=current_year,
        use_date__month=current_month
    )

    # Filter TransactionAccount for the current month
    account_transactions = TransactionAccount.objects.filter(
        txn_date__year=current_year,
        txn_date__month=current_month
    )

    # Combine and sort all transactions by date
    combined_transactions = []
    for t in cash_transactions:
        combined_transactions.append({
            'id': t.id,
            'type': 'cash',
            'date': localtime(t.use_date),
            'side': t.cash_side,
            'amount': t.cash_amount,
            'category': t.cash_cat.cat_type if t.cash_cat else '',
            'content': t.cash_cont,
            'memo': t.memo,
            'photo_url': t.photo.url if t.photo else None,
            'asset_type': t.asset_type, # Use actual asset_type from model
        })
    for t in account_transactions:
        combined_transactions.append({
            'id': t.id,
            'type': 'account',
            'date': localtime(t.txn_date),
            'side': t.txn_side,
            'amount': t.txn_amount,
            'category': t.txn_cat.cat_type if t.txn_cat else '',
            'content': t.txn_cont,
            'memo': '', # Account transactions don't have memo/photo directly
            'photo_url': None,
            'asset_type': t.my_acc.acc_bank, # Use bank name as asset type
        })
    
    # Sort combined transactions by date in descending order
    combined_transactions.sort(key=lambda x: x['date'], reverse=True)
    transactions = combined_transactions # Rename for context

    # Calculate total income for the current month from combined transactions
    monthly_income = sum(
        t['amount'] for t in combined_transactions
        if t['side'] == '수입' or t['side'] == '입금'
    )

    # Calculate total expense for the current month from combined transactions
    monthly_expense = sum(
        t['amount'] for t in combined_transactions
        if t['side'] == '지출' or t['side'] == '출금'
    )

    # Calculate total balance from all bank accounts
    user = User.objects.first() # Placeholder user
    accounts = Account.objects.filter(acc_user_name=user)
    
    latest_transactions_subquery = TransactionAccount.objects.filter(
        my_acc=OuterRef('pk'),
        txn_date__year=current_year,
        txn_date__month=current_month
    ).order_by('-txn_date')
    
    accounts_with_latest_balance = accounts.annotate(
        latest_balance=Subquery(latest_transactions_subquery.values('txn_balance')[:1])
    )
    
    total_bank_balance = accounts_with_latest_balance.aggregate(
        total=Sum('latest_balance')
    )['total'] or 0

    # Calculate current cash balance from the latest cash transaction up to the end of the current month
    # This requires finding the balance of the last transaction that occurred ON or BEFORE the end of the current month
    from datetime import date, timedelta
    import calendar # Import calendar module for month range

    # Get the last day of the current month
    last_day_of_month = calendar.monthrange(current_year, current_month)[1]
    end_of_month_date = date(current_year, current_month, last_day_of_month)

    latest_cash_transaction_up_to_month = TransactionCash.objects.filter(
        use_date__lte=end_of_month_date
    ).order_by('-use_date', '-id').first() # Added -id for consistent ordering if dates are same
    current_cash_balance = latest_cash_transaction_up_to_month.cash_balance if latest_cash_transaction_up_to_month else 0

    # Calculate total assets
    total_assets = current_cash_balance + total_bank_balance
    
    context = {
        'categories': categories,
        'transactions': transactions,
        'monthly_income': monthly_income,
        'monthly_expense': monthly_expense,
        'total_bank_balance': total_bank_balance,
        'total_assets': total_assets,
    }

    return render(request, 'account_book/home.html', context)

#거래내역 저장 뷰 (POST로 넘어온 데이터를 받아서 DB에 저장하는 view 함수를 만든다.)
def save_transaction(request):
    if request.method == "POST":
        cash_side = request.POST.get("cash_side")   # '수입', '지출'
        cash_amount = request.POST.get("amount", "0").replace(",", "")
        try:
            cash_amount = int(cash_amount) if cash_amount else 0
        except ValueError:
            cash_amount = 0
        use_date_str = request.POST.get("date")
        category_name = request.POST.get("category")
        content = request.POST.get("content")
        memo = request.POST.get("memo")
        photo = request.FILES.get("photo") # Get photo from request.FILES
        asset_type = request.POST.get("asset_type") # Retrieve asset_type

        # Parse the date string into a datetime.date object
        parsed_date_obj = None
        parsed_datetime_utc = None # Initialize to None
        if use_date_str:
            try:
                # from datetime import datetime # Already imported at top
                # from django.utils import timezone # Already imported at top
                parsed_date_obj = datetime.strptime(use_date_str, '%Y-%m-%d').date()
                # Convert date object to timezone-aware datetime at midnight UTC
                # This ensures the date is stored consistently without timezone shifts
                parsed_datetime_utc = timezone.make_aware(
                    datetime(parsed_date_obj.year, parsed_date_obj.month, parsed_date_obj.day, 0, 0, 0)
                )
            except ValueError:
                # Handle invalid date format, e.g., log error or return an error response
                pass # For now, just pass, which means parsed_date_obj will remain None

        # --- Calculate new cash balance ---
        last_transaction = TransactionCash.objects.order_by('-use_date').first()
        last_balance = last_transaction.cash_balance if last_transaction else 0

        if cash_side == '수입':
            new_balance = last_balance + cash_amount
        else: # 지출
            new_balance = last_balance - cash_amount
        # -------------------------------------

        # Map front-end '수입'/'지출' to model's 'income'/'expense'
        kind = 'income' if cash_side == '수입' else 'expense'

        category = None
        if category_name:
            try:
                category = AccountBookCategory.objects.get(cat_type=category_name, cat_kind=kind)
            except AccountBookCategory.DoesNotExist:
                # This can happen if JS categories are out of sync with DB
                pass  # Silently ignore if category does not exist

        user = User.objects.first()  # 🔹 로그인 붙이기 전 임시

        TransactionCash.objects.create(
            cash_side=cash_side,
            cash_amount=cash_amount,
            use_date=parsed_datetime_utc, # Use parsed_datetime_utc here
            cash_balance=new_balance,  # Use calculated balance
            cash_cont=content if content else "",
            memo=memo,
            photo=photo,
            cash_cat=category,
            cash_user=user,
            asset_type=asset_type, # Save asset_type
        )
        return redirect("account_book:home")

    return redirect("account_book:home")

@require_POST
def delete_transactions(request):
    ids = []

    # fetch(Ajax) JSON 요청 처리
    if request.headers.get("Content-Type") == "application/json":
        try:
            data = json.loads(request.body)
            ids = data.get("ids", [])
        except json.JSONDecodeError:
            return JsonResponse({"success": False, "error": "잘못된 요청"}, status=400)
    else:
        # 기존 form 전송 처리
        ids = request.POST.getlist('ids[]')

    if ids:
        # TODO: Add user ownership check when user authentication is fully implemented
        TransactionCash.objects.filter(id__in=ids).delete()

        # Ajax 요청인 경우 JSON 응답
        if request.headers.get("Content-Type") == "application/json":
            return JsonResponse({"success": True})

    # 기본적으로는 홈으로 리다이렉트
    return redirect('account_book:home')

@require_POST
def save_bulk_transactions(request):
    if request.headers.get("Content-Type") != "application/json":
        return JsonResponse({"success": False, "error": "Invalid content type"}, status=400)

    try:
        data = json.loads(request.body)
        transactions_data = data.get('transactions', [])
    except json.JSONDecodeError:
        return JsonResponse({"success": False, "error": "Invalid JSON"}, status=400)

    user = User.objects.first()  # 임시 사용자

    try:
        with transaction.atomic():
            # 마지막 잔액 가져오기
            last_transaction = TransactionCash.objects.order_by('-use_date', '-id').first()
            current_balance = last_transaction.cash_balance if last_transaction else 0

            # 날짜 기준으로 먼저 정렬 (잔액 계산의 정확성을 위해)
            transactions_data.sort(key=lambda x: x.get('date', ''))

            for item in transactions_data:
                cash_side = item.get('type')
                cash_amount_str = item.get('amount')
                use_date_str = item.get('date')
                
                # 단건 입력 패널의 유효성 검사와 동일하게, 타입, 금액, 날짜가 모두 있어야 함
                if not all([cash_side, cash_amount_str, use_date_str]):
                    continue

                # Parse the date string into a datetime object
                parsed_date = None
                if use_date_str:
                    try:
                        # Assuming YYYY-MM-DD format from frontend
                        parsed_date = make_aware(datetime.strptime(use_date_str, '%Y-%m-%d'))
                    except ValueError:
                        # Handle invalid date format if necessary, e.g., skip or return error
                        continue # Skip this transaction if date format is invalid

                try:
                    cash_amount = int(cash_amount_str)
                except (ValueError, TypeError):
                    continue # 금액이 유효한 숫자가 아니면 건너뛰기

                category_name = item.get('category')
                content = item.get('content')
                memo = item.get('memo')
                asset_type = item.get('asset') # Retrieve asset_type

                # 잔액 계산
                if cash_side == '수입':
                    current_balance += cash_amount
                else: # 지출
                    current_balance -= cash_amount
                
                kind = 'income' if cash_side == '수입' else 'expense'
                category = None
                if category_name:
                    try:
                        category = AccountBookCategory.objects.filter(cat_type=category_name, cat_kind=kind).first()
                    except AccountBookCategory.DoesNotExist:
                        pass

                TransactionCash.objects.create(
                    cash_user=user,
                    use_date=parsed_date,
                    cash_side=cash_side,
                    cash_cat=category,
                    cash_amount=cash_amount,
                    cash_cont=content,
                    memo=memo,
                    cash_balance=current_balance,
                    asset_type=asset_type # Save asset_type
                )
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)
    

def search_transactions(request):
    user = User.objects.first() # 임시 사용자
    transactions = TransactionCash.objects.filter(cash_user=user).order_by('-use_date')

    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    transaction_type = request.GET.get('type')
    category_id = request.GET.get('category')
    amount_str = request.GET.get('amount')

    year = request.GET.get('year')
    month = request.GET.get('month')

    filters = Q()

    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d') + timedelta(days=1)
        start_date = timezone.make_aware(start_date)
        end_date = timezone.make_aware(end_date)
        filters &= Q(use_date__gte=start_date, use_date__lt=end_date)
    elif year and month:
        filters &= Q(use_date__year=year, use_date__month=month)

    if transaction_type == 'income':
        filters &= Q(cash_side='수입')
    elif transaction_type == 'expense':
        filters &= Q(cash_side='지출')

    if category_id:
        filters &= Q(cash_cat__id=category_id)

    if amount_str:
        try:
            amount = int(amount_str)
            filters &= Q(cash_amount=amount)
        except ValueError:
            pass

    transactions = transactions.filter(filters)

    serialized_transactions = []
    for transaction in transactions:
        serialized_transactions.append({
            'id': transaction.id,
            'use_date': localtime(transaction.use_date).strftime('%Y-%m-%d'),  # ✅ 한국시간 보정
            'cash_side': transaction.cash_side,
            'asset_type': transaction.asset_type,
            'category_name': transaction.cash_cat.cat_type if transaction.cash_cat else '',
            'cash_amount': transaction.cash_amount,
            'cash_cont': transaction.cash_cont,
            'memo': transaction.memo,
            'photo_url': transaction.photo.url if transaction.photo else None,
        })
    
    return JsonResponse({'transactions': serialized_transactions})
def export_transactions_excel(request):
    year = request.GET.get("year")
    month = request.GET.get("month")

    # 쿼리셋 필터링
    transactions = Transaction.objects.all()
    if year and month:
        transactions = transactions.filter(date__year=year, date__month=month)

    # 워크북 만들기
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # 헤더 작성
    headers = ["날짜", "유형", "카테고리", "금액", "메모"]
    ws.append(headers)

    # 데이터 작성
    for t in transactions:
        ws.append([
            t.date.strftime("%Y-%m-%d"),
            t.get_type_display(),  # choices일 경우
            str(t.category) if t.category else "",
            t.amount,
            t.memo or "",
        ])

def export_monthly_transactions(request, year, month):
    # 엑셀 워크북 생성
    wb = openpyxl.Workbook()

    # --- 계좌거래내역 시트 ---
    ws1 = wb.active
    ws1.title = "계좌거래내역"

    ws1.append(["날짜", "거래종류", "금액", "내용", "잔액", "카테고리"])
    account_txns = TransactionAccount.objects.filter(
        txn_date__year=year,
        txn_date__month=month
    )

    for t in account_txns:
        ws1.append([
            t.txn_date.strftime("%Y-%m-%d"),
            t.txn_side,
            t.txn_amount,
            t.txn_cont,
            t.txn_balance,
            t.txn_cat.cat_type if t.txn_cat else ""
        ])

    # --- 현금거래내역 시트 ---
    ws2 = wb.create_sheet(title="현금거래내역")

    ws2.append(["날짜", "구분", "금액", "내용", "메모", "잔액", "카테고리"])
    cash_txns = TransactionCash.objects.filter(
        use_date__year=year,
        use_date__month=month
    )

    for t in cash_txns:
        ws2.append([
            t.use_date.strftime("%Y-%m-%d"),
            t.cash_side,
            t.cash_amount,
            t.cash_cont,
            t.memo,
            t.cash_balance,
            t.cash_cat.cat_type if t.cash_cat else ""
        ])

    # 응답 반환
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"transactions_{year}_{month}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

def index(request):
    transactions = Transaction.objects.all()

    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    type_filter = request.GET.get("type")
    category = request.GET.get("category")

    if start_date:
        transactions = transactions.filter(date__gte=start_date)
    if end_date:
        transactions = transactions.filter(date__lte=end_date)
    if type_filter:
        transactions = transactions.filter(type=type_filter)
    if category:
        transactions = transactions.filter(category=category)

    return render(request, "main/index.html", {"transactions": transactions})
# 검색 카테고리 home화면 
def get_categories(request):
    type_map = {
        "income": "수입",
        "expense": "지출"
    }
    type_param = request.GET.get("type")
    if type_param not in type_map:
        return JsonResponse({"categories": []})

    categories = Category.objects.filter(cat_kind=type_map[type_param])
    data = [{"id": c.id, "name": c.cat_type} for c in categories]
    return JsonResponse({"categories": data})